from __future__ import annotations

import pytest

from benchmark.rule_loader import load_rules


def test_load_real_rules(real_rule_path):
    if not real_rule_path.exists():
        pytest.skip("real rule file not present")
    rs = load_rules(real_rule_path)
    assert len(rs.products) >= 8
    macaw = [p for p in rs.products if p.name.lower().startswith("macaw")]
    assert macaw, "应包含 Macaw 系列 product"
    macaw0 = next(p for p in rs.products if p.name.strip().lower() == "macaw")
    families = [f.family for f in macaw0.filters]
    assert "INAND-MACAW-COMMERCIAL" in families
    assert "INAND-MACAW-INDUSTRIAL" in families

    swift_auto = next(p for p in rs.products if p.name.lower().startswith("swift pro auto"))
    assert any("OTHERS" in f.family for f in swift_auto.filters)
    devs = swift_auto.filters[0].devices
    assert any("SDINFDQ6" in d for d in devs)


def test_step_code(real_rule_path):
    if not real_rule_path.exists():
        pytest.skip("real rule file not present")
    rs = load_rules(real_rule_path)
    assert rs.step_code("7405") in {"SU"}
    assert rs.step_code("7288") == "SB"
    assert rs.step_code("7464") == "QV"
    assert rs.step_code("9999") == "9999"


def test_macaw_amp_separator(real_rule_path):
    """Rule_list 里 Macaw COMMERCIAL 的 STEPID_YIELD 写作 ``7215&7010``,
    应被解析成两个 STEPID."""
    if not real_rule_path.exists():
        pytest.skip("real rule file not present")
    rs = load_rules(real_rule_path)
    macaw = next(p for p in rs.products if p.name.strip().lower() == "macaw")
    commercial = next(f for f in macaw.filters if "COMMERCIAL" in f.family)
    assert set(commercial.stepids) == {"7215", "7010"}


def test_macaw_industrial_inherits_merged_stepid(real_rule_path):
    """Macaw 这个 product table 在 Rule_list 中:
        F4:F5 合并 (STEPID_YIELD = "7215&7010") -> 同时作用于 COMMERCIAL +
        INDUSTRIAL。即使 INDUSTRIAL 那行 STEPID_YIELD 单元格视觉为空,
        也必须继承合并范围内 anchor 的值, 不能解析成 "All"。"""
    if not real_rule_path.exists():
        pytest.skip("real rule file not present")
    rs = load_rules(real_rule_path)
    macaw = next(p for p in rs.products if p.name.strip().lower() == "macaw")
    industrial = next(f for f in macaw.filters if "INDUSTRIAL" in f.family)
    assert set(industrial.stepids) == {"7215", "7010"}, (
        f"INDUSTRIAL 应继承合并 anchor 的 stepids, 实际: {industrial.stepids}"
    )
    assert set(industrial.dieqty) == {"1", "2"}, (
        f"INDUSTRIAL 应继承合并 anchor 的 dieqty, 实际: {industrial.dieqty}"
    )


def test_merged_cells_synthetic(tmp_path):
    """合成一个 Rule_list 验证合并解析: A 列 + 多个数据列都合并 anchor 行的值."""
    import openpyxl

    p = tmp_path / "synth.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product 数据筛选rule"
    ws.append(["Product table name", "FAMILY", "DIEQTY", "MTECH", "Device", "STEPID_YIELD"])
    ws.append(["Foo", "INAND-FOO-A", "1,2", "All", "All", "7010&7215"])
    ws.append([None, "INAND-FOO-B", None, "256Gb", "All", None])
    ws.append(["Bar", "INAND-BAR", "4", "All", "All", "All"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("C2:C3")  # DIEQTY 合并
    ws.merge_cells("F2:F3")  # STEPID_YIELD 合并
    step = wb.create_sheet("Step对应rule")
    step.append(["STEPID_YIELD", "Description"])
    step.append(["7010", "ST_3rd (SG)"])
    step.append(["7215", "MTST1 (SA)"])
    wb.save(p)

    rs = load_rules(p)
    foo = next(r for r in rs.products if r.name == "Foo")
    a = next(f for f in foo.filters if "FOO-A" in f.family)
    b = next(f for f in foo.filters if "FOO-B" in f.family)
    assert set(a.stepids) == {"7010", "7215"}
    assert set(b.stepids) == {"7010", "7215"}, "F 列合并应让 B 也继承"
    assert set(a.dieqty) == set(b.dieqty) == {"1", "2"}, "C 列合并 dieqty 也要继承"
    assert b.mtech == ("256Gb",), "MTECH 没合并, B 自己的值生效"
