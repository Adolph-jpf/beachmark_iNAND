"""Processor + Excel writer 端到端测试."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]

from benchmark.cli import run as cli_run
from benchmark.data_loader import load_yield_data
from benchmark.excel_writer import (
    GREEN,
    RED,
    YELLOW,
    color_for_yield,
    write_excel,
)
from benchmark.goal_loader import load_goal_table
from benchmark.processor import (
    ReportCell,
    build_all_reports,
    format_volume,
    yield_or_none,
)
from benchmark.rule_loader import load_rules
from benchmark.week_calc import select_periods


# --------------------- 颜色规则 ---------------------


def test_color_above_stretch():
    fill, font = color_for_yield(0.999, 0.99, 0.995)
    assert fill == GREEN and font == "FF000000"


def test_color_between_goal_and_stretch():
    fill, font = color_for_yield(0.992, 0.99, 0.995)
    assert fill == GREEN and font == "FFFF0000"


def test_color_within_one_pp():
    fill, font = color_for_yield(0.985, 0.99, 0.995)
    assert fill == YELLOW and font == "FF000000"


def test_color_more_than_one_pp_below_goal():
    fill, font = color_for_yield(0.97, 0.99, 0.995)
    assert fill == RED and font == "FFFFFFFF"


def test_color_no_yield_no_color():
    fill, font = color_for_yield(None, 0.99, 0.995)
    assert fill is None and font is None


def test_color_goal_missing_is_green_yellow():
    fill, font = color_for_yield(0.99, None, None, goal_missing=True)
    assert fill == GREEN and font == YELLOW


# --------------------- % 格式 + A 列合并 ---------------------


def test_yield_cells_use_percent_format(real_data_path, real_rule_path, tmp_path):
    src_path = ROOT / "bachmark SRC.xlsx"
    if not (real_data_path.exists() and real_rule_path.exists() and src_path.exists()):
        pytest.skip("real files not present")
    out = tmp_path / "pct.xlsx"
    cli_run(
        data_path=real_data_path,
        rule_path=real_rule_path,
        output_path=out,
        src_goals_path=src_path,
        today=datetime(2026, 5, 7),
        weeks_count=9,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb[wb.sheetnames[0]]

    # 数值列 (Goal/S-Goal/yield) 的 number_format 应该是百分比格式
    pct_cells = 0
    for row in ws.iter_rows(min_row=2):
        for c in row[4:-1]:  # Goal..最后一列前
            if isinstance(c.value, (int, float)):
                assert "%" in (c.number_format or ""), (
                    f"({c.coordinate}) value={c.value} number_format={c.number_format!r}"
                )
                pct_cells += 1
    assert pct_cells > 5, "应至少有几个百分比格式的数值单元格"


def test_first_column_merged_for_same_group(real_data_path, real_rule_path, tmp_path):
    src_path = ROOT / "bachmark SRC.xlsx"
    if not (real_data_path.exists() and real_rule_path.exists() and src_path.exists()):
        pytest.skip("real files not present")
    out = tmp_path / "merge.xlsx"
    cli_run(
        data_path=real_data_path,
        rule_path=real_rule_path,
        output_path=out,
        src_goals_path=src_path,
        today=datetime(2026, 5, 7),
        weeks_count=9,
    )
    wb = openpyxl.load_workbook(out)
    found_merge = False
    for ws in wb.worksheets:
        for mr in ws.merged_cells.ranges:
            if mr.min_col == 1 and mr.max_col == 1 and mr.max_row > mr.min_row:
                found_merge = True
                break
        if found_merge:
            break
    assert found_merge, "至少应有一处 A 列因相同 group_label 被合并"


def test_write_excel_falls_back_when_locked(synth_yield_text, real_rule_path, tmp_path):
    """模拟"Excel 已打开锁文件"场景: 第一次写 PermissionError, 自动加 _HHMMSS 后缀."""
    from unittest.mock import patch

    if not real_rule_path.exists():
        pytest.skip("real rule file not present")

    # 用合成数据做 reports
    rules = load_rules(real_rule_path)
    df = load_yield_data(synth_yield_text)
    periods = select_periods(df, datetime(2026, 5, 7))
    reports = build_all_reports(df, rules, periods)

    out = tmp_path / "locked.xlsx"
    real_save = openpyxl.Workbook.save
    calls = {"n": 0}

    def fake_save(self, path):
        calls["n"] += 1
        if calls["n"] == 1 and str(path).endswith("locked.xlsx"):
            raise PermissionError("simulate Excel holding the file")
        return real_save(self, path)

    with patch.object(openpyxl.Workbook, "save", fake_save):
        actual = write_excel(reports, out)

    assert actual != out, "第一次失败应自动改名"
    assert actual.name.startswith("locked_") and actual.suffix == ".xlsx"
    assert actual.exists()


def test_macaw_aggregates_across_families(real_data_path, real_rule_path, tmp_path):
    """Macaw 4 个 family 应在同一 sheet 内合并计算; 不再按 family 分行."""
    src_path = ROOT / "bachmark SRC.xlsx"
    if not (real_data_path.exists() and real_rule_path.exists() and src_path.exists()):
        pytest.skip("real files not present")
    df = load_yield_data(real_data_path)
    rules = load_rules(real_rule_path)
    goals = load_goal_table(src_path)
    periods = select_periods(df, datetime(2026, 5, 7))

    macaw_rule = next(r for r in rules.products if r.name.strip().lower() == "macaw")
    assert len(macaw_rule.filters) >= 2

    from benchmark.processor import build_product_report

    rep = build_product_report(df, macaw_rule, rules, periods, goals=goals)
    multi_family_rows = [r for r in rep.rows if len(r.families) >= 2]
    assert multi_family_rows, (
        "Macaw 多个 family 共享同一个 product table, 至少应有一行汇总自 >=2 个 family"
    )


# --------------------- volume 阈值 ---------------------


def test_volume_filter_keeps_high_volume():
    cell = ReportCell(yield_value=0.99, inqty=2000, outqty=1980)
    assert yield_or_none(cell) == pytest.approx(0.99)
    assert format_volume(cell) == 2.0


def test_volume_filter_blanks_low_volume():
    cell = ReportCell(yield_value=0.99, inqty=300, outqty=297)
    assert yield_or_none(cell) is None
    assert format_volume(cell) == "No volume"


def test_volume_filter_handles_none_cell():
    assert yield_or_none(None) is None  # type: ignore[arg-type]
    assert format_volume(None) == "No volume"


# --------------------- 端到端 (合成数据) ---------------------


def test_build_all_reports_synth(synth_yield_text, real_rule_path, tmp_path):
    if not real_rule_path.exists():
        pytest.skip("real rule file not present")
    df = load_yield_data(synth_yield_text)
    rules = load_rules(real_rule_path)
    periods = select_periods(df, datetime(2026, 5, 7))
    reports = build_all_reports(df, rules, periods)
    assert reports
    condor = next(r for r in reports if "condor" in r.name.lower() and "auto" not in r.name.lower())
    assert condor.rows
    out = write_excel(reports, tmp_path / "out.xlsx")
    assert out.exists() and out.stat().st_size > 1024
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames


def test_cli_run_real(real_data_path, real_rule_path, tmp_path):
    if not real_data_path.exists() or not real_rule_path.exists():
        pytest.skip("real files not present")
    out = tmp_path / "real.xlsx"
    meta = cli_run(
        data_path=real_data_path,
        rule_path=real_rule_path,
        output_path=out,
        src_goals_path=None,
        today=datetime(2026, 5, 7),
        weeks_count=9,
        log_file=tmp_path / "run.log",
    )
    assert out.exists()
    assert meta["products"]
    assert len(meta["periods"]["weeks_window"]) == 9
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames
    ws0 = wb[wb.sheetnames[0]]
    headers = [ws0.cell(row=1, column=c).value for c in range(1, ws0.max_column + 1)]
    assert headers[0]
    assert any("Goal" == h for h in headers)
    assert any("S-Goal" == h for h in headers)
    assert any(str(h).startswith("W") for h in headers if h)


def test_cli_run_with_src_goals(real_data_path, real_rule_path, tmp_path):
    """启用 SRC goal table 时, Goal 优先来自 SRC; 缺失行用绿底黄字."""
    src_path = ROOT / "bachmark SRC.xlsx"
    if not (real_data_path.exists() and real_rule_path.exists() and src_path.exists()):
        pytest.skip("real files not present")
    out = tmp_path / "with_src.xlsx"
    meta = cli_run(
        data_path=real_data_path,
        rule_path=real_rule_path,
        output_path=out,
        src_goals_path=src_path,
        today=datetime(2026, 5, 7),
        weeks_count=9,
    )
    assert out.exists()
    assert "missing_goal_rows" in meta

    wb = openpyxl.load_workbook(out)
    found_green_yellow = False
    found_missing_row = False
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            goal_val = row[4].value
            sgoal_val = row[5].value
            is_missing_row = goal_val is None and sgoal_val is None and row[0].value
            if is_missing_row:
                found_missing_row = True
            for c in row[6:-1]:
                fg = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
                font_rgb = c.font.color.rgb if c.font and c.font.color else None
                if fg == GREEN and font_rgb == YELLOW and c.value is not None:
                    found_green_yellow = True
                    break
            if found_green_yellow:
                break
        if found_green_yellow:
            break
    if meta["missing_goal_rows"] > 0:
        assert found_missing_row, "至少应有一行 Goal/S-Goal 都为空 (SRC 中没找到)"
        assert found_green_yellow, "至少应有一个绿底黄字的 yield 单元格"


def test_color_application_in_real_excel(real_data_path, real_rule_path, tmp_path):
    if not real_data_path.exists() or not real_rule_path.exists():
        pytest.skip("real files not present")
    out = tmp_path / "real.xlsx"
    cli_run(
        data_path=real_data_path,
        rule_path=real_rule_path,
        output_path=out,
        src_goals_path=None,
        today=datetime(2026, 5, 7),
        weeks_count=9,
    )
    wb = openpyxl.load_workbook(out)
    found_color = False
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, (int, float)):
                    fg = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
                    if fg in {GREEN, YELLOW, RED}:
                        found_color = True
                        break
            if found_color:
                break
        if found_color:
            break
    assert found_color, "至少应有一个 yield cell 被染色"
