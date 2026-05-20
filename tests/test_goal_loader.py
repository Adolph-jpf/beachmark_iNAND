"""测试 SRC goal 加载器."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from benchmark.goal_loader import GoalEntry, _norm_name, load_goal_table

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="session")
def real_src_path() -> Path:
    return ROOT / "bachmark SRC.xlsx"


def test_norm_name_handles_spaces_and_case():
    assert _norm_name("Swift pro auto") == _norm_name("Swift pro Auto")
    assert _norm_name("Macaw  ") == _norm_name("Macaw")
    assert _norm_name("Condor") == _norm_name("condor")


def test_load_real_src(real_src_path):
    if not real_src_path.exists():
        pytest.skip("real SRC not present")
    gt = load_goal_table(real_src_path)

    assert any(s.lower().startswith("condor") for s in gt.known_tables())

    entry = gt.lookup("condor", "SU", "1", mtech="64Gb")
    assert entry.found
    assert entry.goal == pytest.approx(0.995)
    assert entry.s_goal == pytest.approx(0.995)

    entry = gt.lookup("condor", "SU", "8", mtech="64Gb")
    assert entry.found
    assert entry.goal == pytest.approx(0.982)

    entry = gt.lookup("Swift pro Auto", "SU", "4")
    assert entry.found
    assert entry.goal == pytest.approx(0.982)
    assert entry.s_goal == pytest.approx(0.986)

    entry = gt.lookup("eagle Auto", "SJ", "4")
    assert entry.found and entry.goal == pytest.approx(0.985)
    entry = gt.lookup("eagle Auto", "SK", "8")
    assert entry.found and entry.goal == pytest.approx(0.95)

    miss = gt.lookup("condor", "ZZ", "1")
    assert not miss.found
    assert miss.goal is None and miss.s_goal is None


def test_only_sdss_rows(real_src_path, tmp_path):
    """SwiftPro sheet R2 是 4D JSCC, R3 是 4D SDSS. 我们只取 SDSS."""
    if not real_src_path.exists():
        pytest.skip("real SRC not present")
    gt = load_goal_table(real_src_path)
    entry = gt.lookup("Swift pro", "SB", "4")
    assert entry.found
    assert entry.goal == pytest.approx(0.995)
    assert entry.s_goal == pytest.approx(0.995)


def test_synth_goal_table(tmp_path):
    """用一个手搓 xlsx 测试合并单元格 + Assy 过滤."""
    p = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Demo")
    ws.append(["Demo iNAND", "Config", "Assy", "Goal", "S-Goal", "Q3"])
    ws.append(["BICS3 256Gb MTST SA", "1D", "JSCC", 0.97, 0.98, ""])
    ws.append([None,                  "1D", "SDSS", 0.99, 0.995, ""])
    ws.append([None,                  "2D", "SDSS", 0.985, 0.99, ""])
    ws.append(["BICS3 128Gb MTST SA", "1D", "SDSS", 0.98, 0.99, ""])
    ws.merge_cells("A2:A4")
    wb.save(p)

    gt = load_goal_table(p)
    e = gt.lookup("Demo", "SA", "1", mtech="256Gb")
    assert e.found and e.goal == pytest.approx(0.99)
    e = gt.lookup("Demo", "SA", "2", mtech="256Gb")
    assert e.found and e.goal == pytest.approx(0.985)
    e = gt.lookup("Demo", "SA", "1", mtech="128Gb")
    assert e.found and e.goal == pytest.approx(0.98)
    e = gt.lookup("Demo", "SA", "1", mtech="999Gb")
    assert not e.found, "多个 mtech 候选时入参 mtech 不匹配 -> missing"
    e2 = gt.lookup("Demo", "SA", "9")
    assert not e2.found


def test_lookup_strict_mtech_no_naming_fuzzy(tmp_path):
    """严格 mtech 匹配, 不做命名近似 (``512Tb`` vs ``512Gb`` 视为不同).

    程序不应自作主张把 SRC 的 ``512Tb`` 与原始数据 ``512Gb`` 撮合到一起;
    命名差异请用户自己在 SRC 中对齐。
    """
    p = tmp_path / "single.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SwiftProAuto")
    ws.append(["SwiftPro Auto iNAND", "Config", "Assy", "Goal", "S-Goal"])
    ws.append(["BICS5 512Tb ST SB", "4D", "SDSS", 0.99, 0.995])
    wb.save(p)

    gt = load_goal_table(p)
    e = gt.lookup("SwiftProAuto", "SB", "4", mtech="512Gb")
    assert not e.found, "命名不一致应当 missing, 不做近似"
    e = gt.lookup("SwiftProAuto", "SB", "4", mtech="512Tb")
    assert e.found and e.goal == pytest.approx(0.99)
    e = gt.lookup("SwiftProAuto", "SB", "4", mtech=None)
    assert e.found, "入参 mtech 为空时退到第一条"


def test_lookup_tb_to_gb_equivalence_and_step_alias(tmp_path):
    """业务约定: 1024Gb == 1Tb; step ``SV_-25C`` 可匹配 SRC 的 ``SV``."""
    p = tmp_path / "oberon_like.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Oberon")
    ws.append(["Oberon iNAND", "Config", "Assy", "Goal", "S-Goal"])
    ws.append(["BICS6 1Tb ST1 SB", "4D", "SDSS", 0.995, 0.995])
    ws.append(["BICS6 1Tb ST1 SV", "4D", "SDSS", 0.97, 0.975])
    ws.append(["BICS6 1Tb ST1 SG", "4D", "SDSS", 0.99, 0.995])
    wb.save(p)

    gt = load_goal_table(p)
    # mtech 等价: raw=1024Gb, src=1Tb
    sb = gt.lookup("Oberon", "SB", "4", mtech="1024Gb")
    assert sb.found and sb.goal == pytest.approx(0.995)
    sg = gt.lookup("Oberon", "SG", "4", mtech="1024Gb")
    assert sg.found and sg.goal == pytest.approx(0.99)
    # step 别名: raw=SV_-25C, src=SV
    sv = gt.lookup("Oberon", "SV_-25C", "4", mtech="1024Gb")
    assert sv.found and sv.goal == pytest.approx(0.97)
