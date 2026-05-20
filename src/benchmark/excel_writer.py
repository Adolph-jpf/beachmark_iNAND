"""把 ProductReport 列表写成 Excel 文件 (含颜色规则)."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .processor import (
    ProductReport,
    ReportRow,
    VOLUME_THRESHOLD_K,
    format_volume,
    yield_or_none,
)
from .week_calc import PeriodSelection

GREEN = "FF92D050"
YELLOW = "FFFFFF00"
RED = "FFFF0000"
BLACK = "FF000000"
WHITE = "FFFFFFFF"

THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color=BLACK)
HEADER_FILL = PatternFill("solid", fgColor="FFEDEDED")
ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
PERCENT_FMT = "0.00%"


def color_for_yield(
    yield_value: float | None,
    goal: float | None,
    s_goal: float | None,
    *,
    goal_missing: bool = False,
) -> tuple[str | None, str | None]:
    """根据 Yield 与 Goal/S-Goal 返回 (fill_color, font_color).

    规则 (来自 Screenshot 2026-05-07 004746.jpg):
    * Yield > S-Goal             -> 绿底黑字
    * Goal <= Yield <= S-Goal    -> 绿底红字
    * Goal - Yield <= 1pp        -> 黄底黑字
    * Goal - Yield > 1pp         -> 红底白字

    特殊情况:
    * ``goal_missing=True`` (即 SRC 中没找到该 step 的 goal): 不分级,
      yield cell 用 **绿底黄字** 提示缺 goal.
    * yield 为 None: 不染色, 返回 (None, None).
    """
    if yield_value is None:
        return None, None
    if goal_missing:
        return GREEN, YELLOW
    if goal is None and s_goal is None:
        return None, None
    if s_goal is not None and yield_value > s_goal:
        return GREEN, BLACK
    if goal is not None and s_goal is not None and goal <= yield_value <= s_goal:
        return GREEN, RED
    if goal is not None and yield_value < goal:
        diff = goal - yield_value
        if diff <= 0.01:
            return YELLOW, BLACK
        return RED, WHITE
    return None, None


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\[\]]", "_", name).strip()
    return (cleaned or "Sheet")[:31]


def write_excel(
    reports: list[ProductReport],
    output_path: str | Path,
    *,
    threshold_k: float = VOLUME_THRESHOLD_K,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    for r in reports:
        name = _safe_sheet_name(r.name) or "Product"
        base = name
        i = 2
        while name.lower() in used_names:
            name = f"{base[:28]}_{i}"
            i += 1
        used_names.add(name.lower())
        ws = wb.create_sheet(name)
        _write_sheet(ws, r, threshold_k=threshold_k)

    if not reports:
        ws = wb.create_sheet("EMPTY")
        ws["A1"] = "未生成任何 product report (规则未匹配到任何数据)"

    return _save_with_fallback(wb, output_path)


def _save_with_fallback(wb, output_path: Path) -> Path:
    """正常 ``wb.save``; 文件被 Excel 锁占用时, 自动退到 ``_HHMMSS`` 后缀.

    Windows 下 Excel 打开同名 ``.xlsx`` 会锁文件, ``openpyxl.save`` 会抛
    ``PermissionError``。这里 fallback 一次, 把时间戳追加到 stem 上, 保证
    不会因为用户开着旧报表而中断生成 (旧文件保留, 新文件加时间戳)。
    """
    log = logging.getLogger("benchmark")
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        alt = output_path.with_name(f"{output_path.stem}_{ts}{output_path.suffix}")
        log.warning(
            "%s 被占用 (Excel 打开?), 自动退到 %s", output_path, alt
        )
        wb.save(alt)
        return alt


def _write_sheet(ws, report: ProductReport, *, threshold_k: float) -> None:
    periods = report.periods
    weeks = periods.weeks_window
    title_col = report.name + " iNAND"
    last_week_label = f"{weeks[-1]} Qty (K)" if weeks else "Qty (K)"
    headers = [
        title_col,
        "Step",
        "Config",
        "Assy",
        "Goal",
        "S-Goal",
        periods.quarter,
        periods.month,
        *(_short_week(w) for w in weeks),
        last_week_label,
    ]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN
        cell.border = BORDER

    period_columns = [periods.quarter, periods.month, *weeks]

    for r_idx, row in enumerate(report.rows, start=2):
        a_cell = ws.cell(row=r_idx, column=1, value=row.display_label())
        a_cell.border = BORDER
        a_cell.alignment = ALIGN
        step_cell = ws.cell(row=r_idx, column=2, value=row.step_code)
        step_cell.border = BORDER
        cfg_cell = ws.cell(row=r_idx, column=3, value=row.config)
        cfg_cell.border = BORDER
        ws.cell(row=r_idx, column=4, value=row.assy).border = BORDER

        # 按需求: 第一/二/三列字体加粗
        for c in (a_cell, step_cell, cfg_cell):
            c.font = Font(bold=True, color=BLACK)

        goal_cell = ws.cell(row=r_idx, column=5, value=_num(row.goal))
        goal_cell.border = BORDER
        goal_cell.number_format = PERCENT_FMT
        s_goal_cell = ws.cell(row=r_idx, column=6, value=_num(row.s_goal))
        s_goal_cell.border = BORDER
        s_goal_cell.number_format = PERCENT_FMT

        for offset, period in enumerate(period_columns):
            col = 7 + offset
            cell = row.cells.get(period)
            y = yield_or_none(cell, threshold_k=threshold_k) if cell else None
            xl_cell = ws.cell(row=r_idx, column=col, value=_num(y))
            xl_cell.border = BORDER
            xl_cell.alignment = ALIGN
            xl_cell.number_format = PERCENT_FMT
            fill, font_color = color_for_yield(
                y, row.goal, row.s_goal, goal_missing=row.goal_missing
            )
            if fill:
                xl_cell.fill = PatternFill("solid", fgColor=fill)
            if font_color:
                xl_cell.font = Font(color=font_color)

        last_col = 7 + len(period_columns)
        last_cell_data = row.cells.get(periods.prev_week)
        vol_value = format_volume(last_cell_data, threshold_k=threshold_k)
        xl_cell = ws.cell(row=r_idx, column=last_col, value=vol_value)
        xl_cell.border = BORDER
        xl_cell.alignment = ALIGN
        if isinstance(vol_value, str) and vol_value == "No volume":
            xl_cell.font = Font(italic=True, color="FF7F7F7F")

    _merge_repeated_first_column(ws, report.rows, start_row=2)
    _autosize(ws, headers)
    ws.freeze_panes = "B2"


def _merge_repeated_first_column(ws, rows, *, start_row: int) -> None:
    """连续同 group_label 的行, 合并 A 列单元格."""
    if not rows:
        return
    block_start = start_row
    block_label = rows[0].display_label()
    for i in range(1, len(rows)):
        cur_label = rows[i].display_label()
        if cur_label != block_label:
            end = start_row + i - 1
            if end > block_start:
                ws.merge_cells(start_row=block_start, start_column=1, end_row=end, end_column=1)
            block_start = start_row + i
            block_label = cur_label
    end = start_row + len(rows) - 1
    if end > block_start:
        ws.merge_cells(start_row=block_start, start_column=1, end_row=end, end_column=1)


def _short_week(label: str) -> str:
    """2026FW36 -> W36."""
    m = re.match(r"^\d{4}FW(\d{1,2})$", label)
    return f"W{int(m.group(1))}" if m else label


def _num(value: float | None) -> object:
    """转 0~1 小数; 由 number_format 负责显示成百分比."""
    if value is None:
        return None
    return round(float(value), 6)


def _autosize(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = len(str(h))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=i).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(8, max_len + 2), 28)
