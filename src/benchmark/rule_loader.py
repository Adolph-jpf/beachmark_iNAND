"""加载 Rule_list.xlsx 中的两张表。

Sheet 1: ``Product 数据筛选rule``
    一行规则可能跨多行(Product table name 合并)。返回的数据结构为
    ``ProductRule`` 列表，每个 ProductRule 含若干 ``FamilyFilter``。

Sheet 2: ``Step对应rule``
    一个 STEPID_YIELD -> 描述 的映射。我们再从描述里抓取括号里的简短代码
    (如 SB / QV / SK / S9...) 作为 step 的 ``code``，方便分组显示。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Iterable

import openpyxl

PRODUCT_SHEET_CANDIDATES = ("Product 数据筛选rule", "Product table")
STEP_SHEET_CANDIDATES = ("Step对应rule", "Step rule")

_STEP_CODE_RE = re.compile(r"\(([^()]+)\)\s*$")


@dataclass(frozen=True)
class FamilyFilter:
    """一行 family 级别的过滤规则."""

    family: str
    dieqty: tuple[str, ...]  # 空 tuple 代表 All
    mtech: tuple[str, ...]
    devices: tuple[str, ...]
    stepids: tuple[str, ...]

    def matches_dieqty(self, value: str) -> bool:
        return not self.dieqty or value.strip() in self.dieqty

    def matches_mtech(self, value: str) -> bool:
        return not self.mtech or value.strip() in self.mtech

    def matches_device(self, value: str) -> bool:
        return not self.devices or value.strip() in self.devices

    def matches_stepid(self, value: str) -> bool:
        return not self.stepids or value.strip() in self.stepids


@dataclass(frozen=True)
class ProductRule:
    """一个 product table，可能由多个 family filter 合并产生."""

    name: str
    filters: tuple[FamilyFilter, ...]

    @property
    def families(self) -> tuple[str, ...]:
        return tuple(f.family for f in self.filters)


@dataclass(frozen=True)
class StepInfo:
    stepid: str
    description: str
    code: str  # 提取的简短代码; 若提取失败则取 description


@dataclass(frozen=True)
class RuleSet:
    products: tuple[ProductRule, ...]
    steps: dict[str, StepInfo] = field(default_factory=dict)

    def step_code(self, stepid: str) -> str:
        info = self.steps.get(str(stepid).strip())
        return info.code if info else str(stepid).strip()

    def step_desc(self, stepid: str) -> str:
        info = self.steps.get(str(stepid).strip())
        return info.description if info else ""


def _split_multi(value: object) -> tuple[str, ...]:
    """处理 ``"All"`` / ``"1,2"`` / ``"7215&7010"`` / ``"a\\nb\\nc"`` 等多值字段."""
    if value is None:
        return ()
    s = str(value).strip()
    if not s or s.lower() == "all":
        return ()
    parts: list[str] = []
    for chunk in re.split(r"[\n,&;]", s):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    return tuple(parts)


def _pick_sheet(wb: openpyxl.Workbook, candidates: Iterable[str]) -> str:
    names = wb.sheetnames
    for c in candidates:
        if c in names:
            return c
    # 模糊匹配: 包含关键字
    for name in names:
        for c in candidates:
            if c.split()[0].lower() in name.lower():
                return name
    raise ValueError(f"找不到匹配 sheet, 候选: {candidates}, 实际: {names}")


def load_rules(path: str | Path) -> RuleSet:
    """读取 ``Rule_list.xlsx``."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"规则文件不存在: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)

    product_sheet = _pick_sheet(wb, PRODUCT_SHEET_CANDIDATES)
    step_sheet = _pick_sheet(wb, STEP_SHEET_CANDIDATES)

    products = _load_products(wb[product_sheet])
    steps = _load_steps(wb[step_sheet])
    return RuleSet(products=tuple(products), steps=steps)


def _build_merged_value_getter(ws):
    """返回 ``cell_value(row, col)`` -- 自动解合并单元格的 anchor 值.

    Excel 中合并单元格只有 anchor (左上角) 那一格存值, 其它格 ``value`` 是
    ``None``。但在我们这里, ``Rule_list`` 用合并表示 *该值适用于范围内所有
    家族行* (例如 ``F4:F5`` 合并 = ``STEPID_YIELD=7215&7010`` 同样作用于
    ``INAND-MACAW-INDUSTRIAL``)。这里把任意 (row, col) 自动指回 anchor。
    """
    merged_to_anchor: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        a_row, a_col = mr.min_row, mr.min_col
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (a_row, a_col):
                    merged_to_anchor[(r, c)] = (a_row, a_col)

    def get(row: int, col: int):
        anchor = merged_to_anchor.get((row, col))
        if anchor is not None:
            return ws.cell(row=anchor[0], column=anchor[1]).value
        return ws.cell(row=row, column=col).value

    return get


def _load_products(ws) -> list[ProductRule]:
    if ws.max_row < 2:
        return []

    header = [
        str(ws.cell(row=1, column=c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]

    def col(*keys: str) -> int:
        for i, h in enumerate(header):
            for k in keys:
                if k in h:
                    return i + 1  # 返回 1-based column index
        raise ValueError(f"在 Rule_list 表头里没找到 {keys}: {header}")

    idx_name = col("product table", "table name")
    idx_family = col("family")
    idx_die = col("dieqty")
    idx_mtech = col("mtech")
    idx_device = col("device")
    idx_step = col("stepid")

    cell = _build_merged_value_getter(ws)

    products: list[ProductRule] = []
    current_name: str | None = None
    current_filters: list[FamilyFilter] = []

    for r in range(2, ws.max_row + 1):
        family = cell(r, idx_family)
        if family is None or str(family).strip() == "":
            continue

        # 经过合并填充后, name_cell 一定能拿到当前行所属的 product table name
        name_cell = cell(r, idx_name)
        name = str(name_cell).strip() if name_cell else ""

        if name and name != current_name:
            if current_name is not None and current_filters:
                products.append(
                    ProductRule(name=current_name, filters=tuple(current_filters))
                )
            current_name = name
            current_filters = []

        if current_name is None:
            current_name = str(family).strip()

        f = FamilyFilter(
            family=str(family).strip(),
            dieqty=_split_multi(cell(r, idx_die)),
            mtech=_split_multi(cell(r, idx_mtech)),
            devices=_split_multi(cell(r, idx_device)),
            stepids=_split_multi(cell(r, idx_step)),
        )
        current_filters.append(f)

    if current_name is not None and current_filters:
        products.append(ProductRule(name=current_name, filters=tuple(current_filters)))

    return products


def _load_steps(ws) -> dict[str, StepInfo]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    out: dict[str, StepInfo] = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        stepid = str(row[0]).strip()
        desc = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        m = _STEP_CODE_RE.search(desc)
        code = m.group(1).strip() if m else desc
        out[stepid] = StepInfo(stepid=stepid, description=desc, code=code)
    return out
