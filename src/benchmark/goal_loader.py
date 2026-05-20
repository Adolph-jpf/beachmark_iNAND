"""加载 ``bachmark SRC.xlsx`` 里的 (Goal, S-Goal) 参考表.

每个 sheet 名 = 一个 product table (与 Rule_list 的 ``Product table name`` 对应);
sheet 内列约定:

    A: ``... mtech ... step_code``  (合并单元格, 最后一个 token 是 step code)
    B: ``Config``  (e.g. ``1D``, ``2D``)
    C: ``Assy``    (我们只采 ``SDSS``)
    D: ``Goal``
    E: ``S-Goal``

读取完毕后我们得到一张嵌套查询表 ``GoalTable``。

匹配优先级:
1. (product_table, mtech, step_code, dieqty)
2. (product_table, *, step_code, dieqty)
3. miss -> ``GoalEntry(goal=None, s_goal=None, found=False)``

product_table 名做归一化匹配(忽略空格/大小写)，方便 SRC 里写 ``Swift pro auto``
而 Rule_list 里写 ``Swift pro Auto``。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import openpyxl

ASSY_TARGETS = ("SDSS",)

_MTECH_RE = re.compile(r"\b(\d+(?:G[bB]|T[bB]))\b")
_DIE_RE = re.compile(r"^\s*(\d+)\s*D\s*$", re.IGNORECASE)


@dataclass(frozen=True)
class GoalEntry:
    goal: float | None
    s_goal: float | None
    found: bool
    descriptor: str = ""  # SRC 中 A 列原文, e.g. ``BICS3 256Gb MTST SA``

    @classmethod
    def missing(cls) -> "GoalEntry":
        return cls(goal=None, s_goal=None, found=False, descriptor="")


def _norm_name(name: str | None) -> str:
    if name is None:
        return ""
    return re.sub(r"[\s_]+", "", str(name)).lower()


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if 0 <= v <= 1.5 else None
    s = str(value).strip()
    if not s:
        return None
    if s.endswith("%"):
        try:
            return float(s.rstrip("%")) / 100.0
        except ValueError:
            return None
    try:
        v = float(s)
    except ValueError:
        return None
    return v


def _extract_mtech(text: str) -> str | None:
    m = _MTECH_RE.search(text or "")
    return m.group(1) if m else None


def _extract_step_code(text: str) -> str:
    tokens = (text or "").strip().split()
    return tokens[-1] if tokens else ""


def _norm_mtech(value: str | None) -> str:
    """把 mtech 归一到 ``<gb>gb`` 形式.

    例:
    * ``1024Gb`` -> ``1024gb``
    * ``1Tb`` -> ``1024gb``  (用户约定: 1024Gb 就是 1T)
    """
    if not value:
        return ""
    s = str(value).strip().lower()
    m = re.match(r"^(\d+)\s*([gt])b$", s)
    if not m:
        return s
    n = int(m.group(1))
    unit = m.group(2)
    gb = n * 1024 if unit == "t" else n
    return f"{gb}gb"


def _step_candidates(step_code: str) -> list[str]:
    """step code 候选:
    * 原值
    * 若包含 ``_``，追加前缀 (e.g. ``SV_-25C`` -> ``SV``)
    """
    s = str(step_code or "").strip()
    if not s:
        return []
    out = [s]
    if "_" in s:
        out.append(s.split("_", 1)[0])
    # 去重且保序
    dedup: list[str] = []
    seen: set[str] = set()
    for x in out:
        if x not in seen:
            seen.add(x)
            dedup.append(x)
    return dedup


def _extract_dieqty(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    m = _DIE_RE.match(s)
    if m:
        return m.group(1)
    if s.isdigit():
        return s
    return None


@dataclass
class _BucketEntry:
    mtech: str | None
    goal: float | None
    s_goal: float | None
    descriptor: str


@dataclass
class GoalTable:
    """嵌套表: ``{normalized_table: {(step_code, dieqty): [_BucketEntry, ...]}}``."""

    _data: dict[str, dict[tuple[str, str], list[_BucketEntry]]]
    raw_table_names: dict[str, str]  # normalized -> 原始 sheet name

    def known_tables(self) -> Iterable[str]:
        return self.raw_table_names.values()

    def lookup(
        self,
        product_table: str,
        step_code: str,
        dieqty: str,
        mtech: str | None = None,
    ) -> GoalEntry:
        """查 (product_table, step_code, dieqty, mtech) -> GoalEntry.

        **严格匹配** (筛选靠 ``Rule_list``, 这里只查 Goal/S-Goal):

        1. SRC 中没有 ``(step_code, dieqty)`` 的 entry → ``missing``.
        2. 入参 ``mtech`` 不为空 → 在 bucket 里找 ``mtech`` (大小写不敏感)
           精确匹配, 找到就用, 找不到 → ``missing``. **不做命名近似**
           (例如 ``512Tb`` 与 ``512Gb`` 视为不同, 命名差异请在 SRC 中对齐).
        3. 入参 ``mtech`` 为空 → 退到 ``mtech=None`` 的 entry (即 SRC A 列里
           没标 mtech 的全局 goal); 仍找不到 → 取 bucket 第一条.
        """
        key_table = _norm_name(product_table)
        if key_table not in self._data:
            return GoalEntry.missing()
        die_key = str(dieqty).strip()
        bucket = None
        for sc in _step_candidates(step_code):
            bucket = self._data[key_table].get((sc, die_key))
            if bucket:
                break
        if not bucket:
            return GoalEntry.missing()

        if mtech:
            m_key = _norm_mtech(mtech)
            for e in bucket:
                if e.mtech and _norm_mtech(e.mtech) == m_key:
                    return GoalEntry(
                        goal=e.goal, s_goal=e.s_goal, found=True, descriptor=e.descriptor
                    )
            return GoalEntry.missing()

        for e in bucket:
            if not e.mtech:
                return GoalEntry(
                    goal=e.goal, s_goal=e.s_goal, found=True, descriptor=e.descriptor
                )
        first = bucket[0]
        return GoalEntry(
            goal=first.goal,
            s_goal=first.s_goal,
            found=True,
            descriptor=first.descriptor,
        )


def _resolve_anchor_value(ws, row: int, col: int) -> object:
    """读取合并单元格时返回 anchor 值 (被合并的副单元格在 openpyxl 中是 None)."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    coord = cell.coordinate
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def load_goal_table(path: str | Path) -> GoalTable:
    """读取 ``bachmark SRC.xlsx``."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"SRC goal 文件不存在: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    data: dict[str, dict[tuple[str, str], list[_BucketEntry]]] = {}
    raw_names: dict[str, str] = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        norm = _norm_name(sname)
        if not norm:
            continue
        last_a: str = ""
        bucket: dict[tuple[str, str], list[_BucketEntry]] = {}

        header_row = _find_header_row(ws)
        if header_row is None:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            a_val = _resolve_anchor_value(ws, r, 1)
            b_val = ws.cell(row=r, column=2).value
            c_val = ws.cell(row=r, column=3).value
            d_val = ws.cell(row=r, column=4).value
            e_val = ws.cell(row=r, column=5).value

            a_str = "" if a_val is None else str(a_val).strip()
            if a_str:
                last_a = a_str
            else:
                a_str = last_a

            if not a_str:
                continue
            if b_val is None or str(b_val).strip() == "":
                continue
            if c_val is None or str(c_val).strip().upper() not in ASSY_TARGETS:
                continue

            step_code = _extract_step_code(a_str)
            mtech = _extract_mtech(a_str)
            die = _extract_dieqty(b_val)
            if not step_code or not die:
                continue

            goal = _to_float(d_val)
            s_goal = _to_float(e_val)
            if goal is None and s_goal is None:
                continue

            key = (step_code, die)
            bucket.setdefault(key, []).append(
                _BucketEntry(
                    mtech=mtech, goal=goal, s_goal=s_goal, descriptor=a_str
                )
            )

        if bucket:
            data[norm] = bucket
            raw_names[norm] = sname

    return GoalTable(_data=data, raw_table_names=raw_names)


def _find_header_row(ws) -> int | None:
    """SRC 中部分 sheet (Sapota) 的 header 不在第 1 行, 自动找一下."""
    for r in range(1, min(ws.max_row + 1, 8)):
        row = [
            (ws.cell(row=r, column=c).value or "")
            for c in range(1, min(ws.max_column + 1, 6))
        ]
        joined = " ".join(str(x).lower() for x in row)
        if "config" in joined and "goal" in joined and "assy" in joined:
            return r
    return 1
